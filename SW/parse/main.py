"""
Sensor Packet Parser GUI
─────────────────────────
입력:  시리얼(Arduino에 'p' 전송) 또는 .bin 파일
출력:  .xlsx (Excel) 및 .bin 파일
구조:  sensor_data.h를 자동 파싱하여 어떤 구조체든 동적 지원
"""

import os, sys, struct, threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import serial
import serial.tools.list_ports
import openpyxl

from header_parser import parse_header, StructDef

# ── 경로 설정 ──
HERE = os.path.dirname(os.path.abspath(__file__))
HEADER_PATH = os.path.join(HERE, 'sensor_data.h')

# ── 헤더 파싱 ──
STRUCTS = parse_header(HEADER_PATH)  # {pkt_id: StructDef}

# id → 예상 패킷 크기 테이블 (헤더에서 자동 생성)
EXPECTED_SIZE = {sd.pkt_id: sd.size for sd in STRUCTS.values()}
KNOWN_IDS = set(EXPECTED_SIZE.keys())
SYNC = 0xAA


# ═══════════════════════════════════════════════
#  바이트 스트림 → 패킷 리스트 파싱
# ═══════════════════════════════════════════════
def parse_buffer(buf: bytearray):
    """buf에서 가능한 모든 패킷을 추출하여 (pkt_id, raw_bytes) 리스트를 반환.
       buf는 소비된 만큼 앞에서 잘린다.
       패킷 구조: [0xAA] [id] [len] [payload...]"""
    packets = []

    while len(buf) >= 3:
        # 1) sync 바이트 찾기 — 0xAA가 아니면 스킵
        if buf[0] != SYNC:
            buf.pop(0)
            continue

        pkt_id  = buf[1]
        pkt_len = buf[2]

        # 2) len 유효성: 최소 헤더(3) 이상
        if pkt_len < 3:
            buf.pop(0)
            continue

        # 3) 알려진 id인지 확인
        if pkt_id not in KNOWN_IDS:
            buf.pop(0)
            continue

        # 4) 알려진 id면 len이 예상 크기와 일치하는지 확인
        expected = EXPECTED_SIZE.get(pkt_id)
        if expected is not None and pkt_len != expected:
            buf.pop(0)
            continue

        # 5) 전체 패킷이 아직 안 들어왔으면 대기
        if len(buf) < pkt_len:
            break

        raw = bytes(buf[:pkt_len])
        del buf[:pkt_len]
        packets.append((pkt_id, raw))

    return packets


# ═══════════════════════════════════════════════
#  Hex 덤프 유틸리티
# ═══════════════════════════════════════════════
def hexdump(data: bytes, bytes_per_line: int = 16) -> str:
    """바이트 데이터를 hex dump 문자열로 변환 (offset | hex | ascii)"""
    lines = []
    for i in range(0, len(data), bytes_per_line):
        chunk = data[i:i + bytes_per_line]
        hex_part = ' '.join(f'{b:02X}' for b in chunk)
        ascii_part = ''.join(chr(b) if 32 <= b < 127 else '.' for b in chunk)
        lines.append(f'{i:08X}  {hex_part:<{bytes_per_line * 3}}  |{ascii_part}|')
    return '\n'.join(lines)


# ═══════════════════════════════════════════════
#  시리얼 수신 스레드
# ═══════════════════════════════════════════════
class SerialReader(threading.Thread):
    def __init__(self, port, baud, on_packet, on_raw, on_serial_text):
        super().__init__(daemon=True)
        self.port = port
        self.baud = baud
        self.on_packet = on_packet          # callback(pkt_id, decoded_dict)
        self.on_raw    = on_raw             # callback(raw_bytes) — bin 저장용
        self.on_serial_text = on_serial_text # callback(chunk_bytes) — 시리얼 모니터용
        self.running = False
        self.ser = None

    def run(self):
        self.ser = serial.Serial(self.port, self.baud, timeout=0.1)
        self.running = True
        buf = bytearray()

        while self.running:
            chunk = self.ser.read(1024)
            if chunk:
                buf.extend(chunk)
                self.on_serial_text(chunk)  # 수신된 raw chunk를 시리얼 모니터로

            for pkt_id, raw in parse_buffer(buf):
                self.on_raw(raw)
                sd = STRUCTS.get(pkt_id)
                if sd:
                    try:
                        decoded = sd.decode(raw)
                        decoded['_type'] = sd.name.upper()
                        self.on_packet(pkt_id, decoded)
                    except Exception as e:
                        self.on_packet(pkt_id, {'_type': 'ERROR', '_msg': str(e)})
                else:
                    self.on_packet(pkt_id, {'_type': f'UNKNOWN(id={pkt_id})'})

    def send(self, data: bytes):
        if self.ser and self.ser.is_open:
            self.ser.write(data)

    def stop(self):
        self.running = False
        if self.ser and self.ser.is_open:
            self.ser.close()


# ═══════════════════════════════════════════════
#  GUI
# ═══════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Sensor Packet Parser")
        self.geometry("1050x750")
        self.reader = None
        self.raw_bytes = bytearray()   # bin 저장용
        self.all_rows = []             # (pkt_id, decoded_dict) 리스트

        self._build_ui()
        self.refresh_ports()
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    # ── UI 구성 ──
    def _build_ui(self):
        # ── 상단: 시리얼 제어 ──
        frm_serial = ttk.LabelFrame(self, text="Serial", padding=5)
        frm_serial.pack(fill=tk.X, padx=5, pady=(5, 2))

        ttk.Label(frm_serial, text="Port:").pack(side=tk.LEFT)
        self.combo_port = ttk.Combobox(frm_serial, width=12, state="readonly")
        self.combo_port.pack(side=tk.LEFT, padx=3)

        ttk.Label(frm_serial, text="Baud:").pack(side=tk.LEFT)
        self.entry_baud = ttk.Entry(frm_serial, width=8)
        self.entry_baud.insert(0, "115200")
        self.entry_baud.pack(side=tk.LEFT, padx=3)

        self.btn_refresh = ttk.Button(frm_serial, text="Refresh", command=self.refresh_ports)
        self.btn_refresh.pack(side=tk.LEFT, padx=3)

        self.btn_connect = ttk.Button(frm_serial, text="Connect", command=self.toggle_connect)
        self.btn_connect.pack(side=tk.LEFT, padx=3)

        self.btn_send_p = ttk.Button(frm_serial, text="Send 'p'", command=self.send_p, state=tk.DISABLED)
        self.btn_send_p.pack(side=tk.LEFT, padx=3)

        # ── 상단: 파일 입출력 ──
        frm_file = ttk.LabelFrame(self, text="File I/O", padding=5)
        frm_file.pack(fill=tk.X, padx=5, pady=2)

        self.btn_open_bin = ttk.Button(frm_file, text="Open .bin", command=self.open_bin)
        self.btn_open_bin.pack(side=tk.LEFT, padx=3)

        self.btn_save_bin = ttk.Button(frm_file, text="Save .bin", command=self.save_bin)
        self.btn_save_bin.pack(side=tk.LEFT, padx=3)

        self.btn_save_xlsx = ttk.Button(frm_file, text="Save .xlsx", command=self.save_xlsx)
        self.btn_save_xlsx.pack(side=tk.LEFT, padx=3)

        self.btn_clear = ttk.Button(frm_file, text="Clear All", command=self.clear_all)
        self.btn_clear.pack(side=tk.LEFT, padx=3)

        self.lbl_status = ttk.Label(frm_file, text="Packets: 0")
        self.lbl_status.pack(side=tk.RIGHT, padx=10)

        # ── 메인 영역: Notebook 탭 ──
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # ── 탭 1: 파싱 테이블 ──
        frm_table = ttk.Frame(self.notebook)
        self.notebook.add(frm_table, text="  Parsed Data  ")

        all_fields = []
        seen = set()
        for sd in STRUCTS.values():
            for f in sd.fields:
                if f not in seen:
                    seen.add(f)
                    all_fields.append(f)

        self.columns = ['type'] + all_fields
        self.tree = ttk.Treeview(frm_table, columns=self.columns, show="headings", height=18)

        for col in self.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=90, anchor=tk.CENTER)
        self.tree.column('type', width=70)

        vsb = ttk.Scrollbar(frm_table, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(frm_table, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vsb.grid(row=0, column=1, sticky='ns')
        hsb.grid(row=1, column=0, sticky='ew')
        frm_table.rowconfigure(0, weight=1)
        frm_table.columnconfigure(0, weight=1)

        # ── 탭 2: 시리얼 모니터 ──
        frm_monitor = ttk.Frame(self.notebook)
        self.notebook.add(frm_monitor, text="  Serial Monitor  ")

        # 표시 모드 선택
        frm_mon_ctrl = ttk.Frame(frm_monitor)
        frm_mon_ctrl.pack(fill=tk.X, padx=3, pady=3)

        ttk.Label(frm_mon_ctrl, text="Display:").pack(side=tk.LEFT)
        self.serial_mode = tk.StringVar(value="hex")
        ttk.Radiobutton(frm_mon_ctrl, text="HEX", variable=self.serial_mode, value="hex").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(frm_mon_ctrl, text="ASCII", variable=self.serial_mode, value="ascii").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(frm_mon_ctrl, text="Both", variable=self.serial_mode, value="both").pack(side=tk.LEFT, padx=5)

        self.var_autoscroll = tk.BooleanVar(value=True)
        ttk.Checkbutton(frm_mon_ctrl, text="Auto-scroll", variable=self.var_autoscroll).pack(side=tk.LEFT, padx=10)

        btn_clear_mon = ttk.Button(frm_mon_ctrl, text="Clear", command=self._clear_serial_monitor)
        btn_clear_mon.pack(side=tk.RIGHT, padx=3)

        self.txt_serial = tk.Text(frm_monitor, wrap=tk.NONE, font=("Consolas", 10),
                                  bg="#1e1e1e", fg="#d4d4d4", insertbackground="#d4d4d4")
        self.txt_serial.config(state=tk.DISABLED)

        mon_vsb = ttk.Scrollbar(frm_monitor, orient=tk.VERTICAL, command=self.txt_serial.yview)
        mon_hsb = ttk.Scrollbar(frm_monitor, orient=tk.HORIZONTAL, command=self.txt_serial.xview)
        self.txt_serial.configure(yscrollcommand=mon_vsb.set, xscrollcommand=mon_hsb.set)

        self.txt_serial.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        mon_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        mon_hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # ── 탭 3: Hex 뷰어 ──
        frm_hex = ttk.Frame(self.notebook)
        self.notebook.add(frm_hex, text="  Hex Viewer  ")

        frm_hex_ctrl = ttk.Frame(frm_hex)
        frm_hex_ctrl.pack(fill=tk.X, padx=3, pady=3)

        btn_refresh_hex = ttk.Button(frm_hex_ctrl, text="Refresh", command=self._refresh_hex_view)
        btn_refresh_hex.pack(side=tk.LEFT, padx=3)

        self.lbl_hex_info = ttk.Label(frm_hex_ctrl, text="0 bytes")
        self.lbl_hex_info.pack(side=tk.LEFT, padx=10)

        self.txt_hex = tk.Text(frm_hex, wrap=tk.NONE, font=("Consolas", 10),
                               bg="#1e1e1e", fg="#d4d4d4", insertbackground="#d4d4d4")
        self.txt_hex.config(state=tk.DISABLED)

        # hex 뷰어 색상 태그
        self.txt_hex.tag_configure("offset", foreground="#569cd6")   # 파란색: 오프셋
        self.txt_hex.tag_configure("header", foreground="#ff6b6b")   # 빨간색: 패킷 헤더(id,len)
        self.txt_hex.tag_configure("data",   foreground="#98c379")   # 초록색: 패킷 데이터
        self.txt_hex.tag_configure("ascii",  foreground="#808080")   # 회색: ASCII

        hex_vsb = ttk.Scrollbar(frm_hex, orient=tk.VERTICAL, command=self.txt_hex.yview)
        hex_hsb = ttk.Scrollbar(frm_hex, orient=tk.HORIZONTAL, command=self.txt_hex.xview)
        self.txt_hex.configure(yscrollcommand=hex_vsb.set, xscrollcommand=hex_hsb.set)

        self.txt_hex.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        hex_vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hex_hsb.pack(side=tk.BOTTOM, fill=tk.X)

    # ── 포트 ──
    def refresh_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.combo_port['values'] = ports
        if ports:
            self.combo_port.current(0)

    # ── 시리얼 연결 ──
    def toggle_connect(self):
        if self.reader and self.reader.running:
            self.reader.stop()
            self.reader = None
            self.btn_connect.config(text="Connect")
            self.btn_send_p.config(state=tk.DISABLED)
        else:
            port = self.combo_port.get()
            baud = int(self.entry_baud.get())
            if not port:
                return
            try:
                self.reader = SerialReader(port, baud,
                                           self._on_packet, self._on_raw,
                                           self._on_serial_chunk)
                self.reader.start()
                self.btn_connect.config(text="Disconnect")
                self.btn_send_p.config(state=tk.NORMAL)
            except Exception as e:
                messagebox.showerror("Serial Error", str(e))

    def send_p(self):
        if self.reader:
            self.reader.send(b'p')

    # ── 콜백 (스레드 → UI) ──
    def _on_raw(self, raw: bytes):
        self.raw_bytes.extend(raw)

    def _on_packet(self, pkt_id, decoded: dict):
        self.after(0, self._insert_row, pkt_id, decoded)

    def _on_serial_chunk(self, chunk: bytes):
        self.after(0, self._append_serial_monitor, chunk)

    # ── 시리얼 모니터 ──
    def _append_serial_monitor(self, chunk: bytes):
        mode = self.serial_mode.get()

        if mode == "hex":
            text = ' '.join(f'{b:02X}' for b in chunk) + ' '
        elif mode == "ascii":
            text = chunk.decode('ascii', errors='replace')
        else:  # both
            hex_str = ' '.join(f'{b:02X}' for b in chunk)
            asc_str = ''.join(chr(b) if 32 <= b < 127 else '.' for b in chunk)
            text = f'{hex_str}  |{asc_str}|\n'

        self.txt_serial.config(state=tk.NORMAL)
        self.txt_serial.insert(tk.END, text)
        # 최대 10000줄 유지
        line_count = int(self.txt_serial.index('end-1c').split('.')[0])
        if line_count > 10000:
            self.txt_serial.delete('1.0', f'{line_count - 10000}.0')
        self.txt_serial.config(state=tk.DISABLED)

        if self.var_autoscroll.get():
            self.txt_serial.see(tk.END)

    def _clear_serial_monitor(self):
        self.txt_serial.config(state=tk.NORMAL)
        self.txt_serial.delete('1.0', tk.END)
        self.txt_serial.config(state=tk.DISABLED)

    # ── Hex 뷰어 (색상으로 패킷 구조 표시) ──
    def _refresh_hex_view(self):
        data = bytes(self.raw_bytes)
        self.lbl_hex_info.config(text=f"{len(data)} bytes")

        self.txt_hex.config(state=tk.NORMAL)
        self.txt_hex.delete('1.0', tk.END)

        if not data:
            self.txt_hex.config(state=tk.DISABLED)
            return

        # 패킷 경계 맵 생성: 각 바이트가 어느 역할인지 표시
        # 'H' = 헤더(id,len), 'D' = 데이터, ' ' = 알 수 없음
        role = [' '] * len(data)
        pos = 0
        while pos + 1 < len(data):
            pkt_id = data[pos]
            pkt_len = data[pos + 1]
            if pkt_len < 2 or pos + pkt_len > len(data):
                pos += 1
                continue
            expected = EXPECTED_SIZE.get(pkt_id)
            if expected is not None and pkt_len != expected:
                pos += 1
                continue
            role[pos]     = 'H'  # id
            role[pos + 1] = 'H'  # len
            for j in range(pos + 2, pos + pkt_len):
                role[j] = 'D'
            pos += pkt_len

        # 렌더링
        BPL = 16  # bytes per line
        for i in range(0, len(data), BPL):
            chunk = data[i:i + BPL]
            chunk_role = role[i:i + BPL]

            # 오프셋
            self.txt_hex.insert(tk.END, f'{i:08X}  ', 'offset')

            # Hex 바이트 (역할별 색상)
            for j, b in enumerate(chunk):
                r = chunk_role[j]
                tag = 'header' if r == 'H' else 'data' if r == 'D' else ''
                self.txt_hex.insert(tk.END, f'{b:02X} ', tag if tag else ())

            # 16바이트 미만이면 패딩
            if len(chunk) < BPL:
                pad = '   ' * (BPL - len(chunk))
                self.txt_hex.insert(tk.END, pad)

            # ASCII
            asc = ''.join(chr(b) if 32 <= b < 127 else '.' for b in chunk)
            self.txt_hex.insert(tk.END, f' |{asc}|\n', 'ascii')

        self.txt_hex.config(state=tk.DISABLED)

    # ── 파싱 테이블 ──
    def _insert_row(self, pkt_id, decoded: dict):
        self.all_rows.append((pkt_id, decoded))

        values = []
        for col in self.columns:
            if col == 'type':
                values.append(decoded.get('_type', '?'))
            else:
                v = decoded.get(col, '')
                if isinstance(v, float):
                    values.append(f'{v:.4f}')
                else:
                    values.append(str(v) if v != '' else '')
        self.tree.insert("", tk.END, values=values)
        self.tree.yview_moveto(1.0)
        self.lbl_status.config(text=f"Packets: {len(self.all_rows)}")

    # ── .bin 파일 열기 ──
    def open_bin(self):
        path = filedialog.askopenfilename(
            filetypes=[("Binary files", "*.bin"), ("All files", "*.*")])
        if not path:
            return

        with open(path, 'rb') as f:
            data = f.read()

        self.clear_all()
        self.raw_bytes.extend(data)

        buf = bytearray(data)
        for pkt_id, raw in parse_buffer(buf):
            sd = STRUCTS.get(pkt_id)
            if sd:
                try:
                    decoded = sd.decode(raw)
                    decoded['_type'] = sd.name.upper()
                    self._insert_row(pkt_id, decoded)
                except Exception as e:
                    self._insert_row(pkt_id, {'_type': 'ERROR', '_msg': str(e)})
            else:
                self._insert_row(pkt_id, {'_type': f'UNKNOWN(id={pkt_id})'})

        # bin 열면 자동으로 Hex 뷰어도 갱신
        self._refresh_hex_view()

        messagebox.showinfo("Open .bin",
                            f"Loaded {len(self.all_rows)} packets from\n{path}")

    # ── .bin 저장 ──
    def save_bin(self):
        if not self.raw_bytes:
            messagebox.showwarning("Save .bin", "No data to save.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".bin",
            filetypes=[("Binary files", "*.bin")])
        if not path:
            return
        with open(path, 'wb') as f:
            f.write(self.raw_bytes)
        messagebox.showinfo("Save .bin", f"Saved {len(self.raw_bytes)} bytes\n{path}")

    # ── .xlsx 저장 ──
    def save_xlsx(self):
        if not self.all_rows:
            messagebox.showwarning("Save .xlsx", "No data to save.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return

        wb = openpyxl.Workbook()

        # 센서 타입별로 시트 분리
        sheets = {}  # pkt_id -> worksheet
        for pkt_id, decoded in self.all_rows:
            sd = STRUCTS.get(pkt_id)
            if sd is None:
                continue
            if pkt_id not in sheets:
                ws = wb.create_sheet(title=sd.name.upper())
                ws.append(sd.fields)
                sheets[pkt_id] = ws
            ws = sheets[pkt_id]
            row = [decoded.get(f, '') for f in sd.fields]
            ws.append(row)

        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        wb.save(path)
        messagebox.showinfo("Save .xlsx", f"Saved {len(self.all_rows)} packets\n{path}")

    # ── 초기화 ──
    def clear_all(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.all_rows.clear()
        self.raw_bytes.clear()
        self.lbl_status.config(text="Packets: 0")
        self._clear_serial_monitor()
        # Hex 뷰어도 초기화
        self.txt_hex.config(state=tk.NORMAL)
        self.txt_hex.delete('1.0', tk.END)
        self.txt_hex.config(state=tk.DISABLED)
        self.lbl_hex_info.config(text="0 bytes")

    def on_close(self):
        if self.reader:
            self.reader.stop()
        self.destroy()


if __name__ == "__main__":
    App().mainloop()
